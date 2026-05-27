import os
import re
import csv
from datetime import datetime, timedelta

def parse_vietnamese_number(text):
    """
    Quy đổi số viết bằng chữ tiếng Việt sang số nguyên.
    """
    num_map = {
        "một": 1, "hai": 2, "ba": 3, "bốn": 4, "năm": 5,
        "sáu": 6, "bảy": 7, "tám": 8, "chín": 9, "mười": 10
    }
    text_clean = text.strip().lower()
    return num_map.get(text_clean, None)

def parse_relative_date(relative_str, base_date):
    """
    Ước lượng ngày đánh giá tuyệt đối từ chuỗi thời gian tương đối của Google Maps.
    Ví dụ: "3 năm trước" -> base_date - 3 năm.
    """
    if not relative_str or not isinstance(relative_str, str):
        return None
    
    # Loại bỏ tiền tố không cần thiết
    clean_str = relative_str.replace("Thời gian chỉnh sửa:", "").strip().lower()
    
    # Thay thế chữ viết thường thành số tương ứng
    words = clean_str.split()
    for i, w in enumerate(words):
        num_val = parse_vietnamese_number(w)
        if num_val is not None:
            words[i] = str(num_val)
    clean_str = " ".join(words)
    
    # Tìm mẫu "số + đơn vị thời gian"
    match = re.search(r'(\d+)\s+(năm|tháng|tuần|ngày|giờ|phút)', clean_str)
    if match:
        val = int(match.group(1))
        unit = match.group(2)
        if unit == "năm":
            return (base_date - timedelta(days=val * 365)).strftime('%Y-%m-%d')
        elif unit == "tháng":
            return (base_date - timedelta(days=val * 30)).strftime('%Y-%m-%d')
        elif unit == "tuần":
            return (base_date - timedelta(days=val * 7)).strftime('%Y-%m-%d')
        elif unit == "ngày":
            return (base_date - timedelta(days=val)).strftime('%Y-%m-%d')
        elif unit == "giờ":
            return (base_date - timedelta(hours=val)).strftime('%Y-%m-%d')
        elif unit == "phút":
            return (base_date - timedelta(minutes=val)).strftime('%Y-%m-%d')
            
    # Trường hợp không có số cụ thể (ví dụ: "một năm trước" đã thay thành "1 năm trước", 
    # nhưng phòng hờ các cụm từ khác)
    if "năm trước" in clean_str:
        return (base_date - timedelta(days=365)).strftime('%Y-%m-%d')
    if "tháng trước" in clean_str:
        return (base_date - timedelta(days=30)).strftime('%Y-%m-%d')
    if "tuần trước" in clean_str:
        return (base_date - timedelta(days=7)).strftime('%Y-%m-%d')
    if "ngày trước" in clean_str:
        return (base_date - timedelta(days=1)).strftime('%Y-%m-%d')
    if "giờ trước" in clean_str:
        return (base_date - timedelta(hours=1)).strftime('%Y-%m-%d')
        
    # Nếu không parse được, trả về None hoặc ngày base_date mặc định
    return None

def clean_google_reviews(review_dir, output_path=None):
    """
    Đọc, gộp và làm sạch toàn bộ dữ liệu Google Maps Reviews của các thương hiệu.
    """
    all_reviews = []
    base_default_date = datetime(2026, 5, 22) # Baseline dựa trên thời điểm crawl trong metadata
    
    if not os.path.exists(review_dir):
        raise FileNotFoundError(f"Thư mục review không tồn tại: {review_dir}")
        
    for file_name in os.listdir(review_dir):
        if not file_name.endswith(".csv"):
            continue
            
        file_path = os.path.join(review_dir, file_name)
        
        # Xác định thương hiệu dựa trên tên file
        brand_name = "Chưa rõ"
        if "hoa sen home" in file_name.lower():
            brand_name = "Hoa Sen Home"
        elif "rita vo" in file_name.lower() or "rita võ" in file_name.lower():
            brand_name = "Rita Võ"
        elif "viglacera" in file_name.lower():
            brand_name = "Viglacera"
            
        with open(file_path, mode='r', encoding='utf-8-sig', errors='ignore') as f:
            reader = csv.DictReader(f)
            for row in reader:
                # 1. Trích xuất thông tin cơ bản
                user_name = row.get('user_name', '').strip()
                review_text = row.get('review_text', '').strip()
                place_name = row.get('place_name', '').strip()
                
                # 2. Chuẩn hóa rating (Ví dụ: "5 sao" -> 5)
                rating_raw = row.get('rating', '')
                rating = None
                if rating_raw:
                    digit_match = re.search(r'\d+', rating_raw)
                    if digit_match:
                        rating = int(digit_match.group())
                
                # 3. Xử lý thời gian
                # Nếu có cột stage2_finished_at, dùng nó làm base_date để tính ngược
                base_date = base_default_date
                finished_at = row.get('stage2_finished_at', '')
                if finished_at:
                    try:
                        # Thử parse datetime từ định dạng '2026-05-22T23:26:58'
                        base_date = datetime.fromisoformat(finished_at.split('.')[0])
                    except ValueError:
                        pass
                
                review_date_raw = row.get('review_date', '')
                review_date = parse_relative_date(review_date_raw, base_date)
                
                # Nếu không ước lượng được ngày, gán mặc định là ngày base_date
                if not review_date:
                    review_date = base_date.strftime('%Y-%m-%d')
                    
                all_reviews.append({
                    "brand": brand_name,
                    "place_name": place_name,
                    "user_name": user_name,
                    "rating": rating,
                    "review_text": review_text,
                    "review_date": review_date,
                    "review_date_raw": review_date_raw
                })
                
    # Ghi ra file đầu ra nếu có chỉ định
    if output_path and all_reviews:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        headers = ["brand", "place_name", "user_name", "rating", "review_text", "review_date", "review_date_raw"]
        with open(output_path, mode='w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(all_reviews)
            
    return all_reviews

def clean_catalog_products(catalog_dir, output_path=None):
    """
    Làm sạch dữ liệu Catalog Sản phẩm từ tệp balanced_products.csv hoặc gộp từ các file lẻ.
    """
    # Ưu tiên sử dụng balanced_products.csv vì chứa đầy đủ sản phẩm và đã được định dạng
    source_file = os.path.join(catalog_dir, "balanced_products.csv")
    if not os.path.exists(source_file):
        # Nếu không có file balanced_products.csv, thử tìm các file lẻ để gộp
        print(f"Không tìm thấy {source_file}, đang tìm kiếm các tệp sản phẩm riêng lẻ...")
        individual_files = ["hoa_sen_home_products.csv", "rita_vo_products.csv", "viglacera_products.csv"]
        all_products = []
        
        for f_name in individual_files:
            f_path = os.path.join(catalog_dir, f_name)
            if os.path.exists(f_path):
                with open(f_path, mode='r', encoding='utf-8-sig', errors='ignore') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        all_products.append(row)
    else:
        all_products = []
        with open(source_file, mode='r', encoding='utf-8-sig', errors='ignore') as f:
            reader = csv.DictReader(f)
            for row in reader:
                all_products.append(row)
                
    cleaned_products = []
    
    for row in all_products:
        brand = row.get('brand', '').strip()
        product_name = row.get('product_name', '').strip()
        sku = row.get('sku', '').strip()
        product_url = row.get('product_url', '').strip()
        image_url = row.get('image_url', '').strip()
        description = row.get('description', '').strip()
        
        # 1. Làm sạch giá bán
        regular_price = None
        sale_price = None
        
        for p_key in ['regular_price_vnd', 'regular_price']:
            p_val = row.get(p_key, '')
            if p_val:
                try:
                    regular_price = float(re.sub(r'[^\d.]', '', p_val))
                except ValueError:
                    pass
                    
        for p_key in ['sale_price_vnd', 'sale_price']:
            p_val = row.get(p_key, '')
            if p_val:
                try:
                    sale_price = float(re.sub(r'[^\d.]', '', p_val))
                except ValueError:
                    pass
                    
        # Nếu sale_price trống nhưng có regular_price, gán sale_price = regular_price
        if regular_price is not None and sale_price is None:
            sale_price = regular_price
        elif sale_price is not None and regular_price is None:
            regular_price = sale_price
            
        # Tính tỷ lệ giảm giá (%)
        discount_rate = 0.0
        if regular_price and regular_price > 0 and sale_price:
            discount_rate = round((regular_price - sale_price) / regular_price * 100, 2)
            
        # 2. Xử lý phân cấp danh mục (category_path)
        category_path = row.get('category_path', '').strip()
        category_l1 = "Chưa phân loại"
        category_l2 = "Chưa phân loại"
        
        if category_path:
            # Tách danh mục theo ký tự phân tách '|' hoặc '>' hoặc '/'
            parts = [p.strip() for p in re.split(r'[\|>\/]', category_path) if p.strip()]
            if len(parts) >= 1:
                category_l1 = parts[0]
            if len(parts) >= 2:
                category_l2 = parts[1]
                
        # 3. Chuẩn hóa tình trạng kho hàng
        availability_raw = row.get('availability', '').strip().lower()
        availability = True
        if availability_raw in ['out_of_stock', 'outofstock', 'false', '0', 'hết hàng']:
            availability = False
            
        cleaned_products.append({
            "brand": brand,
            "product_name": product_name,
            "sku": sku,
            "category_l1": category_l1,
            "category_l2": category_l2,
            "regular_price_vnd": regular_price,
            "sale_price_vnd": sale_price,
            "discount_rate": discount_rate,
            "availability": availability,
            "product_url": product_url,
            "image_url": image_url,
            "description": description
        })
        
    # Ghi ra file đầu ra nếu có chỉ định
    if output_path and cleaned_products:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        headers = [
            "brand", "product_name", "sku", "category_l1", "category_l2", 
            "regular_price_vnd", "sale_price_vnd", "discount_rate", "availability", 
            "product_url", "image_url", "description"
        ]
        with open(output_path, mode='w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(cleaned_products)
            
    return cleaned_products

def clean_facebook_posts(fb_dir, output_path=None):
    """
    Đọc, gộp và làm sạch dữ liệu bài viết Facebook từ các file Excel xlsx.
    """
    # Sử dụng openpyxl để đọc Excel
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Cần cài đặt thư viện 'openpyxl' để xử lý file Excel Facebook.")
        
    all_posts = []
    
    if not os.path.exists(fb_dir):
        raise FileNotFoundError(f"Thư mục Facebook thô không tồn tại: {fb_dir}")
        
    for file_name in os.listdir(fb_dir):
        if not file_name.endswith(".xlsx") or file_name.startswith("~$"):
            continue
            
        file_path = os.path.join(fb_dir, file_name)
        
        # Xác định thương hiệu dựa trên tên file
        brand_name = "Chưa rõ"
        if "hoa sen home" in file_name.lower():
            brand_name = "Hoa Sen Home"
        elif "rita vo" in file_name.lower() or "rita võ" in file_name.lower():
            brand_name = "Rita Võ"
        elif "viglacera" in file_name.lower():
            brand_name = "Viglacera"
            
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if 'Data' not in wb.sheetnames:
            print(f"Cảnh báo: File {file_name} không có sheet 'Data'. Bỏ qua.")
            continue
            
        sheet = wb['Data']
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
            
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        
        for row_vals in rows[1:]:
            # Bỏ qua dòng trống
            if all(v is None for v in row_vals):
                continue
                
            row_dict = dict(zip(headers, row_vals))
            
            # Trích xuất các trường thông tin cơ bản
            post_id = str(row_dict.get('postId', '') or '').strip()
            url = str(row_dict.get('url', '') or '').strip()
            text = str(row_dict.get('text', '') or '').strip()
            
            # Xử lý thời gian bài đăng
            post_time_raw = row_dict.get('time', '')
            post_time = None
            if post_time_raw:
                post_time_str = str(post_time_raw).split('.')[0].replace('Z', '')
                try:
                    # Chuyển đổi định dạng ISO '2026-05-23T12:01:38'
                    post_time = datetime.fromisoformat(post_time_str).strftime('%Y-%m-%d %H:%M:%S')
                except ValueError:
                    post_time = str(post_time_raw)
                    
            # Trích xuất dạng bài viết Video
            is_video_raw = str(row_dict.get('isVideo', '') or '').lower().strip()
            is_video = True if is_video_raw in ['true', '1', 'yes'] else False
            
            # Trích xuất và ép kiểu các chỉ số tương tác
            def clean_num(val):
                if val is None or val == '':
                    return 0
                try:
                    return int(float(str(val).replace(',', '')))
                except ValueError:
                    return 0
                    
            likes = clean_num(row_dict.get('likes'))
            comments = clean_num(row_dict.get('comments'))
            shares = clean_num(row_dict.get('shares'))
            views_count = clean_num(row_dict.get('viewsCount'))
            
            # Các phản ứng chi tiết
            react_like = clean_num(row_dict.get('reactionLikeCount'))
            react_love = clean_num(row_dict.get('reactionLoveCount'))
            react_haha = clean_num(row_dict.get('reactionHahaCount'))
            react_wow = clean_num(row_dict.get('reactionWowCount'))
            react_sad = clean_num(row_dict.get('reactionSadCount'))
            react_angry = clean_num(row_dict.get('reactionAngryCount'))
            react_care = clean_num(row_dict.get('reactionCareCount'))
            
            # Tính tổng tương tác
            total_engagement = likes + comments + shares
            
            # Tính chỉ số cảm xúc tích cực vs tiêu cực
            positive_reacts = react_love + react_haha + react_wow + react_care
            negative_reacts = react_sad + react_angry
            
            all_posts.append({
                "brand": brand_name,
                "post_id": post_id,
                "url": url,
                "post_time": post_time,
                "text": text,
                "is_video": is_video,
                "likes": likes,
                "comments": comments,
                "shares": shares,
                "views_count": views_count,
                "total_engagement": total_engagement,
                "react_like": react_like,
                "react_love": react_love,
                "react_haha": react_haha,
                "react_wow": react_wow,
                "react_sad": react_sad,
                "react_angry": react_angry,
                "react_care": react_care,
                "positive_reacts": positive_reacts,
                "negative_reacts": negative_reacts
            })
            
    # Ghi ra file đầu ra nếu có chỉ định
    if output_path and all_posts:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        headers = [
            "brand", "post_id", "url", "post_time", "text", "is_video", 
            "likes", "comments", "shares", "views_count", "total_engagement",
            "react_like", "react_love", "react_haha", "react_wow", "react_sad", "react_angry", "react_care",
            "positive_reacts", "negative_reacts"
        ]
        with open(output_path, mode='w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(all_posts)
            
    return all_posts
